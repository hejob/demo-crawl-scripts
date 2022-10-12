from selenium import webdriver
# from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
# from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import csv
import openpyxl


chrome_options = webdriver.chrome.options.Options()
chrome_options.headless = True
# driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options) 
driver = webdriver.Chrome('./chromedriver', options=chrome_options)
# How much time should Selenium wait until an element is able to interact
driver.implicitly_wait(5)

def extract_video_info_from_detail(url):
    driver.get(url)
    print(driver.title)

    element = WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.XPATH, "//h1"))
    )
    title = element.text
    element2 = driver.find_element('xpath', '//*[@id="v_upinfo"]/div[2]/div[1]/a[1]')
    name = element2.text
    element3 = driver.find_element('xpath', '//*[@id="v_upinfo"]/div[2]/div[1]/a[1]')
    mid = element3.get_attribute('href')
    return (title, mid, name)

def extract_info_from_bvids(bvids):
    results = []
    for bvid in bvids:
        url = f'https://www.bilibili.com/video/{bvid}'
        title, mid, name = extract_video_info_from_detail(url)
        results.append( {
            'bvid': bvid,
            'title': title,
            'mid': mid,
            'name': name
        } )
    return results

def output_info_to_csv(results, file_path):
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        infowriter = csv.DictWriter(csvfile, fieldnames=['bvid', 'title', 'mid', 'name'])
        infowriter.writeheader()
        infowriter.writerows(results)

def output_info_to_excel(results, file_path):
    wb = openpyxl.Workbook()

    sheet = wb.active

    fieldnames=['bvid', 'title', 'mid', 'name']
    row = 1
    for (j, field) in enumerate(fieldnames):
        col = 1 + j
        c = sheet.cell(row = row, column = col)
        c.number_format = openpyxl.styles.numbers.FORMAT_GENERAL
        c.value = field
    
    for i, json_row in enumerate(results):
        row = 2 + i
        for j, field_name in enumerate(fieldnames):
            col = j + 1
            c = sheet.cell(row = row, column = col)
            c.number_format = openpyxl.styles.numbers.FORMAT_GENERAL
            c.value = json_row[field_name]

    wb.save(file_path)

### main
if __name__ == '__main__':
    bvids = [ 'BV1c4411e77t', 'BV1aN4y1A7j1' ]
    try:
        results = extract_info_from_bvids(bvids)

        print(results)

        output_info_to_csv(results, "info.csv")
        output_info_to_excel(results, "info.xlsx")
    finally:
        driver.quit()    
