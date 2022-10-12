import enum
from urllib.request import Request, urlopen
import io
import gzip
import json
import csv
import openpyxl

def get_html_content(url):
    """
        Gets content of a url in gzip compressed response
    """
    req = Request(url)
    req.add_header('user-agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.49')
    req.add_header('referer', url)
    req.add_header('accept', 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng')
    req.add_header('Accept-Encoding', 'gzip')

    content = urlopen(req).read()

    gzip_file = io.BytesIO(content)
    with gzip.open(gzip_file, 'rt', encoding='utf-8') as f:
        content = f.read()
    return content

def extract_video_info_from_json_detail(json_data):
    data = json.loads(json_data)
    title = data['data']['title']
    owner = data['data']['owner']
    mid = owner['mid']
    name = owner['name']
    # print(title, mid, name)
    return (title, mid, name)

def extract_info_from_bvids(bvids):
    results = []
    for bvid in bvids:
        url = f'http://api.bilibili.com/x/web-interface/view/?bvid={bvid}'
        json_data = get_html_content(url)
        title, mid, name = extract_video_info_from_json_detail(json_data)
        results.append( {
            'bvid': bvid,
            'title': title,
            'mid': mid,
            'name': name
        } )
    return results

def output_info_to_csv(results, file_path):
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        infowriter = csv.DictWriter(csvfile, fieldnames=['bvid', 'title', 'mid', 'name'])
        infowriter.writeheader()
        infowriter.writerows(results)

def output_info_to_excel(results, file_path):
    wb = openpyxl.Workbook()

    sheet = wb.active

    fieldnames=['bvid', 'title', 'mid', 'name']
    row = 1
    for (j, field) in enumerate(fieldnames):
        col = 1 + j
        c = sheet.cell(row = row, column = col)
        c.number_format = openpyxl.styles.numbers.FORMAT_GENERAL
        c.value = field
    
    for i, json_row in enumerate(results):
        row = 2 + i
        for j, field_name in enumerate(fieldnames):
            col = j + 1
            c = sheet.cell(row = row, column = col)
            c.number_format = openpyxl.styles.numbers.FORMAT_GENERAL
            c.value = json_row[field_name]

    wb.save(file_path)

### main
if __name__ == '__main__':
    bvids = [ 'BV1c4411e77t', 'BV1aN4y1A7j1' ]
    results = extract_info_from_bvids(bvids)

    print(results)

    output_info_to_csv(results, "info.csv")
    output_info_to_excel(results, "info.xlsx")
 

### ---- TODO ----
### - [ ] exceptions
